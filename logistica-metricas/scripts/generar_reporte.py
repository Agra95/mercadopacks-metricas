#!/usr/bin/env python3
"""
generar_reporte.py
-------------------
Toma el export crudo de envios del sistema de logistica (.xls o .xlsx) y
genera un reporte de metricas (.xlsx) con:
  - Resumen general (KPIs)
  - Ranking de choferes
  - Desglose por zona
  - Detalle fila por fila con las columnas calculadas

Las definiciones de negocio (que cuenta como "entrega efectiva", "cancelado",
el corte horario, etc.) NO estan hardcodeadas sueltas en el codigo: viven en
la seccion CONFIG de abajo, que refleja lo acordado en reglas_de_negocio.md.
Si una regla de negocio cambia, se actualiza reglas_de_negocio.md primero y
despues esta seccion.

Uso:
    python3 generar_reporte.py <ruta_al_archivo_crudo.xls> [--salida ruta_salida.xlsx]

Ejemplo:
    python3 generar_reporte.py ../datos_crudos/listado_envios_20260901160852.xls
"""

import argparse
import subprocess
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIG — refleja reglas_de_negocio.md. Actualizar ahi primero, aca despues.
# ============================================================================

ESTADOS_ENTREGA_EFECTIVA = {"Entregado", "Entregado 2DA visita"}
ESTADOS_CANCELADO = {"Cancelado", "Rechazado por el comprador"}
HORA_CORTE = 21  # "antes de las 21hs" -> hora(Fecha estado) < HORA_CORTE
COLUMNA_TIMESTAMP_ENTREGA = "Fecha estado"
MIN_ENVIOS_PARA_RANKING_CONFIABLE = 5  # por debajo de esto, se marca la fila

FUENTE_PRINCIPAL = "Arial"


# ============================================================================
# CARGA Y LIMPIEZA
# ============================================================================

def cargar_export(ruta: Path) -> pd.DataFrame:
    """
    Carga el archivo crudo del sistema de logistica.

    El archivo trae unas filas de "Filtros Aplicados" antes del header real,
    asi que se detecta automaticamente donde empieza la tabla buscando la
    fila que contiene 'ID (Interno)' en la primera columna, en vez de asumir
    un numero fijo de filas a saltear (el numero de filas de filtros puede
    cambiar segun cuantos filtros se hayan aplicado al descargar).

    Los .xls viejos que entrega este sistema a veces vienen con el BIFF
    corrupto (falla xlrd con 'Workbook corruption') aunque Excel/LibreOffice
    los abren sin problema. Si falla la lectura directa, se convierte primero
    a .xlsx con LibreOffice y se reintenta.
    """
    sufijo = ruta.suffix.lower()

    def _leer_crudo(path, engine=None):
        return pd.read_excel(path, header=None, engine=engine)

    try:
        if sufijo == ".xls":
            df_crudo = _leer_crudo(ruta, engine="xlrd")
        else:
            df_crudo = _leer_crudo(ruta, engine="openpyxl")
    except Exception as e:
        print(f"[aviso] No se pudo leer el archivo directamente ({e}).")
        print("[aviso] Reintentando via conversion con LibreOffice...")
        df_crudo = _leer_crudo(_convertir_a_xlsx(ruta), engine="openpyxl")

    # Buscar la fila de header real (contiene 'ID (Interno)' en la col 0)
    fila_header = None
    for i in range(min(20, len(df_crudo))):
        valor = df_crudo.iloc[i, 0]
        if isinstance(valor, str) and valor.strip() == "ID (Interno)":
            fila_header = i
            break

    if fila_header is None:
        raise ValueError(
            "No se encontro la fila de encabezado ('ID (Interno)') en las "
            "primeras 20 filas del archivo. Revisar si el formato del "
            "export cambio."
        )

    columnas = df_crudo.iloc[fila_header].tolist()
    df = df_crudo.iloc[fila_header + 1:].copy()
    df.columns = [str(c).strip() if pd.notna(c) else f"col_{i}" for i, c in enumerate(columnas)]

    # Descartar filas sin ID (Interno): el export trae una fila de totales al
    # final (suma de la columna Precio) que no es un envio real. No alcanza
    # con dropna(how="all") porque esa fila si tiene un valor cargado.
    # Ver reglas_de_negocio.md, seccion 6.
    df = df[df["ID (Interno)"].notna()]

    return df.reset_index(drop=True)


def _convertir_a_xlsx(ruta: Path) -> Path:
    """Convierte un .xls problematico a .xlsx usando LibreOffice headless."""
    tmp_dir = Path(tempfile.mkdtemp())
    resultado = subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(tmp_dir), str(ruta)],
        capture_output=True, text=True, timeout=120,
    )
    if resultado.returncode != 0:
        raise RuntimeError(f"Fallo la conversion con LibreOffice: {resultado.stderr}")
    convertido = tmp_dir / (ruta.stem + ".xlsx")
    if not convertido.exists():
        raise RuntimeError(f"La conversion no genero el archivo esperado en {convertido}")
    return convertido


def procesar(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega las columnas calculadas segun reglas_de_negocio.md."""
    df = df.copy()

    df["Estado"] = df["Estado"].astype(str).str.strip()
    df["Cadete"] = df["Cadete"].fillna("").astype(str).str.strip()
    df["Zona"] = df["Zona"].fillna("Sin zona").astype(str).str.strip().replace("", "Sin zona")

    df["_fecha_estado_dt"] = pd.to_datetime(
        df[COLUMNA_TIMESTAMP_ENTREGA], format="%d/%m/%Y %H:%M", errors="coerce"
    )

    df["es_entrega_efectiva"] = df["Estado"].isin(ESTADOS_ENTREGA_EFECTIVA)
    df["es_cancelado"] = df["Estado"].isin(ESTADOS_CANCELADO)
    df["tiene_chofer"] = df["Cadete"] != ""

    df["es_antes_de_corte"] = (
        df["es_entrega_efectiva"]
        & df["_fecha_estado_dt"].notna()
        & (df["_fecha_estado_dt"].dt.hour < HORA_CORTE)
    )

    return df


# ============================================================================
# CALCULO DE METRICAS
# ============================================================================

def calcular_resumen(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    entregas_efectivas = int(df["es_entrega_efectiva"].sum())
    cancelados = int(df["es_cancelado"].sum())
    antes_de_corte = int(df["es_antes_de_corte"].sum())
    otros = total - entregas_efectivas - cancelados

    filas = [
        ("Total de envios en el periodo", total, None),
        ("Entregas efectivas", entregas_efectivas, _pct(entregas_efectivas, total)),
        (f"Entregas antes de las {HORA_CORTE}hs (sobre entregas efectivas)", antes_de_corte, _pct(antes_de_corte, entregas_efectivas)),
        (f"Entregas antes de las {HORA_CORTE}hs (sobre total de envios)", antes_de_corte, _pct(antes_de_corte, total)),
        ("Cancelados / rechazados", cancelados, _pct(cancelados, total)),
        ("Otros estados (en curso / pendiente de definir)", otros, _pct(otros, total)),
    ]
    return pd.DataFrame(filas, columns=["Metrica", "Cantidad", "% "])


def calcular_por_chofer(df: pd.DataFrame) -> pd.DataFrame:
    con_chofer = df[df["tiene_chofer"]]
    sin_chofer_total = int((~df["tiene_chofer"]).sum())

    agg = con_chofer.groupby("Cadete").agg(
        total_asignados=("Estado", "count"),
        entregas_efectivas=("es_entrega_efectiva", "sum"),
        antes_de_corte=("es_antes_de_corte", "sum"),
        cancelados=("es_cancelado", "sum"),
    ).reset_index()

    agg["% efectividad"] = (agg["entregas_efectivas"] / agg["total_asignados"] * 100).round(1)
    agg["% antes de corte (s/ entregas)"] = (
        agg["antes_de_corte"] / agg["entregas_efectivas"].replace(0, np.nan) * 100
    ).round(1)
    agg["muestra_chica"] = agg["total_asignados"] < MIN_ENVIOS_PARA_RANKING_CONFIABLE

    agg = agg.sort_values("% efectividad", ascending=False).reset_index(drop=True)

    fila_sin_chofer = pd.DataFrame([{
        "Cadete": "(sin chofer asignado)",
        "total_asignados": sin_chofer_total,
        "entregas_efectivas": pd.NA,
        "antes_de_corte": pd.NA,
        "cancelados": pd.NA,
        "% efectividad": pd.NA,
        "% antes de corte (s/ entregas)": pd.NA,
        "muestra_chica": pd.NA,
    }])

    return pd.concat([agg, fila_sin_chofer], ignore_index=True)


def calcular_por_zona(df: pd.DataFrame) -> pd.DataFrame:
    agg = df.groupby("Zona").agg(
        total=("Estado", "count"),
        entregas_efectivas=("es_entrega_efectiva", "sum"),
        antes_de_corte=("es_antes_de_corte", "sum"),
        cancelados=("es_cancelado", "sum"),
    ).reset_index()
    agg["% efectividad"] = (agg["entregas_efectivas"] / agg["total"] * 100).round(1)
    agg["% cancelados"] = (agg["cancelados"] / agg["total"] * 100).round(1)
    return agg.sort_values("total", ascending=False).reset_index(drop=True)


def _pct(parte, total):
    if not total:
        return None
    return round(parte / total * 100, 1)


# ============================================================================
# ESCRITURA DEL REPORTE .xlsx
# ============================================================================

def escribir_reporte(ruta_salida: Path, resumen, por_chofer, por_zona, detalle):
    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    _volcar_df(ws_resumen, resumen)

    ws_chofer = wb.create_sheet("Por Chofer")
    _volcar_df(ws_chofer, por_chofer)

    ws_zona = wb.create_sheet("Por Zona")
    _volcar_df(ws_zona, por_zona)

    columnas_detalle = [
        "ID (Interno)", "Número Tracking", "Estado", "Cadete", "Zona",
        "Método de envío", "Fecha estado", "es_entrega_efectiva",
        "es_antes_de_corte", "es_cancelado",
    ]
    columnas_detalle = [c for c in columnas_detalle if c in detalle.columns]
    ws_detalle = wb.create_sheet("Detalle")
    _volcar_df(ws_detalle, detalle[columnas_detalle])

    wb.save(ruta_salida)


def _volcar_df(ws, df: pd.DataFrame):
    header_font = Font(name=FUENTE_PRINCIPAL, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    body_font = Font(name=FUENTE_PRINCIPAL)

    for j, col in enumerate(df.columns, start=1):
        celda = ws.cell(row=1, column=j, value=str(col))
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = Alignment(horizontal="center")

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            celda = ws.cell(row=i, column=j, value=val)
            celda.font = body_font

    ws.freeze_panes = "A2"
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    for j, col in enumerate(df.columns, start=1):
        largo = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(j)].width = min(max(largo + 2, 10), 45)


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description="Genera el reporte de metricas de envios.")
    parser.add_argument("archivo_crudo", type=str, help="Ruta al export crudo (.xls o .xlsx)")
    parser.add_argument("--salida", type=str, default=None, help="Ruta del .xlsx de salida")
    args = parser.parse_args()

    ruta_entrada = Path(args.archivo_crudo)
    if not ruta_entrada.exists():
        sys.exit(f"No se encontro el archivo: {ruta_entrada}")

    if args.salida:
        ruta_salida = Path(args.salida)
    else:
        fecha = pd.Timestamp.now().strftime("%Y-%m-%d")
        ruta_salida = ruta_entrada.parent.parent / "reportes" / f"reporte_{fecha}.xlsx"
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    print(f"Leyendo {ruta_entrada} ...")
    df_crudo = cargar_export(ruta_entrada)
    print(f"  {len(df_crudo)} envios encontrados.")

    df = procesar(df_crudo)

    resumen = calcular_resumen(df)
    por_chofer = calcular_por_chofer(df)
    por_zona = calcular_por_zona(df)

    escribir_reporte(ruta_salida, resumen, por_chofer, por_zona, df)

    print()
    print("=== RESUMEN ===")
    print(resumen.to_string(index=False))
    print()
    print(f"Reporte completo guardado en: {ruta_salida}")


if __name__ == "__main__":
    main()
